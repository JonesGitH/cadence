"""Settings routes — storage, business info, calendars, Graph auth, import/export."""
import io
import json
import logging
import os
import shutil
from datetime import datetime

from flask import render_template, request, redirect, url_for, flash, jsonify, send_file

from app import app
import config as _config
import graph_auth as _graph_auth
from database import (
    get_db, get_setting, set_setting, get_settings_batch,
    get_all_calendars, upsert_calendars, backup_database,
    password_is_set, password_expires_in_days,
)
from outlook import discover_calendars
from helpers import SERVICES

log = logging.getLogger(__name__)

_SETTINGS_DEFAULTS = {
    'hourly_rate':          '0',
    'business_name':        'Your Name',
    'business_title':       '',
    'business_email':       'your@email.com',
    'business_phone':       '(555) 555-5555',
    'business_address':     '',
    'business_city':        '',
    'business_state':       '',
    'business_zip':         '',
    'venmo_handle':         '',
    'graph_client_id':      '',
    'graph_timezone':       'Central Standard Time',
    'idle_timeout_minutes': '30',
}


@app.route('/settings')
def settings():
    cfg   = get_settings_batch(tuple(_SETTINGS_DEFAULTS))
    s     = {k: cfg.get(k, v) for k, v in _SETTINGS_DEFAULTS.items()}
    paths = _config.load()
    return render_template('settings.html', s=s, calendars=get_all_calendars(), paths=paths,
                           graph_connected=_graph_auth.is_connected(),
                           pw_set=password_is_set(),
                           pw_expires_in=password_expires_in_days())


@app.route('/settings/backup', methods=['POST'])
def backup_db():
    try:
        path = backup_database()
        log.info('Database backed up to %s', path)
        flash(f'Backup saved: {path}', 'success')
    except Exception as e:
        log.error('Backup failed: %s', e)
        flash(f'Backup failed: {e}', 'error')
    return redirect(url_for('settings'))


@app.route('/settings/student-template.xlsx')
def student_import_template():
    """Download a pre-formatted Excel template for bulk student import."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students'

    COLS = [
        ('name',            'Name *',              22, 'Required. Student first + last name.'),
        ('initials',        'Initials *',           10, 'Required. 2-3 letters used to match Outlook events (e.g. AJ).'),
        ('email',           'Email',                26, 'Student or family email for invoice delivery.'),
        ('phone',           'Phone',                16, 'Contact phone number.'),
        ('school',          'School',               22, 'School name.'),
        ('grade',           'Grade',                10, 'K · 1st–12th · College · Other'),
        ('birthday',        'Birthday',             14, 'YYYY-MM-DD'),
        ('diagnosis',       'Diagnosis',            28, 'Primary diagnosis / learning profile.'),
        ('services',        'Services',             40,
         'Comma-separated keys: ' + ', '.join(k for k, _ in SERVICES)),
        ('services_other',  'Services (Other)',     24, 'Free-text description when "other" is included in Services.'),
        ('start_date',      'Start Date',           14, 'YYYY-MM-DD'),
        ('end_date',        'End Date',             14, 'YYYY-MM-DD (leave blank if ongoing)'),
        ('test_date',       'Test Date',            14, 'YYYY-MM-DD'),
        ('parent1_name',    'Parent 1 Name',        22, 'Primary parent / guardian name.'),
        ('parent2_name',    'Parent 2 Name',        22, 'Second parent / guardian name.'),
        ('parent_address',  'Street Address',       26, 'Parent 1 billing street address.'),
        ('parent_city',     'City',                 18, 'Parent 1 city.'),
        ('parent_state',    'State',                8,  '2-letter state code (e.g. TX)'),
        ('parent_zip',      'Zip',                  10, 'Parent 1 zip.'),
        ('parent2_address', 'Street Address (P2)',  26, 'Parent 2 billing street address.'),
        ('parent2_city',    'City (P2)',             18, 'Parent 2 city.'),
        ('parent2_state',   'State (P2)',            8,  '2-letter state code (e.g. TX)'),
        ('parent2_zip',     'Zip (P2)',              10, 'Parent 2 zip.'),
        ('parent2_phone',        'Phone (P2)',            16, 'Parent 2 contact phone number.'),
        ('parent2_email',        'Email (P2)',            26, 'Parent 2 email address.'),
        ('bill_to_parent',       'Bill To',              10, 'Who receives the invoice: 1 (Parent 1), 2 (Parent 2), or custom. Defaults to 1.'),
        ('bill_to_custom_name',  'Custom Bill-To Name',  24, 'Full name when Bill To = custom.'),
        ('bill_to_custom_addr',  'Custom Bill-To Street',26, 'Street address when Bill To = custom.'),
        ('bill_to_custom_city',  'Custom Bill-To City',  18, 'City when Bill To = custom.'),
        ('bill_to_custom_state', 'Custom Bill-To State',  8, '2-letter state when Bill To = custom.'),
        ('bill_to_custom_zip',   'Custom Bill-To Zip',   10, 'Zip when Bill To = custom.'),
        ('intake_complete', 'Intake Complete',      16, 'YES or NO'),
        ('roi_complete',    'ROI Complete',         14, 'YES or NO'),
        ('hourly_rate',     'Per-Session Rate',     16, 'Leave blank to use default rate from Settings.'),
        ('notes',           'Notes',                40, 'Internal notes (not printed on invoices).'),
    ]

    hdr_fill  = PatternFill('solid', fgColor='1E2D3D')
    req_fill  = PatternFill('solid', fgColor='2563EB')
    note_fill = PatternFill('solid', fgColor='F0F4FF')
    hdr_font  = Font(bold=True, color='FFFFFF', size=11)
    req_font  = Font(bold=True, color='FFFFFF', size=11)
    note_font = Font(italic=True, color='374151', size=9)
    thin      = Side(style='thin', color='D1D5DB')
    thin_border = Border(left=thin, right=thin, bottom=thin, top=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=False)
    wrap   = Alignment(horizontal='left',   vertical='top',    wrap_text=True)

    for col_idx, (key, label, width, _note) in enumerate(COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        is_req = label.endswith('*')
        cell.fill      = req_fill if is_req else hdr_fill
        cell.font      = req_font if is_req else hdr_font
        cell.alignment = center
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for col_idx, (key, label, width, note) in enumerate(COLS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=note)
        cell.fill      = note_fill
        cell.font      = note_font
        cell.alignment = wrap
        cell.border    = thin_border
    ws.row_dimensions[2].height = 28

    ws.freeze_panes = 'A3'

    GRADES_LIST = ['K', '1st', '2nd', '3rd', '4th', '5th', '6th', '7th',
                   '8th', '9th', '10th', '11th', '12th', 'College', 'Other']
    grade_col    = next(i for i, (k, *_) in enumerate(COLS, 1) if k == 'grade')
    intake_col   = next(i for i, (k, *_) in enumerate(COLS, 1) if k == 'intake_complete')
    roi_col      = next(i for i, (k, *_) in enumerate(COLS, 1) if k == 'roi_complete')
    bill_to_col  = next(i for i, (k, *_) in enumerate(COLS, 1) if k == 'bill_to_parent')

    dv_grade = DataValidation(type='list', formula1=f'"{",".join(GRADES_LIST)}"',
                               showDropDown=False, showErrorMessage=True,
                               errorTitle='Invalid grade',
                               error='Choose a value from the dropdown list.')
    dv_grade.sqref = f'{get_column_letter(grade_col)}3:{get_column_letter(grade_col)}1000'
    ws.add_data_validation(dv_grade)

    for c in (intake_col, roi_col):
        dv_yn = DataValidation(type='list', formula1='"YES,NO"',
                                showDropDown=False, showErrorMessage=True,
                                errorTitle='Invalid value', error='Enter YES or NO.')
        dv_yn.sqref = f'{get_column_letter(c)}3:{get_column_letter(c)}1000'
        ws.add_data_validation(dv_yn)

    dv_bill = DataValidation(type='list', formula1='"1,2,custom"',
                              showDropDown=False, showErrorMessage=True,
                              errorTitle='Invalid value', error='Enter 1, 2, or custom.')
    dv_bill.sqref = f'{get_column_letter(bill_to_col)}3:{get_column_letter(bill_to_col)}1000'
    ws.add_data_validation(dv_bill)

    SAMPLE = {
        'name': 'Jane Smith', 'initials': 'JS', 'email': 'parent@example.com',
        'phone': '555-123-4567', 'school': 'Lincoln Elementary', 'grade': '4th',
        'birthday': '2015-03-22', 'diagnosis': 'Dyslexia',
        'services': 'reading,reading_comprehension',
        'services_other': '', 'start_date': '2025-09-01', 'end_date': '',
        'test_date': '2025-08-15', 'parent1_name': 'Mary Smith',
        'parent2_name': 'John Smith', 'parent_address': '123 Oak St',
        'parent_city': 'Austin', 'parent_state': 'TX', 'parent_zip': '78701',
        'parent2_address': '', 'parent2_city': '', 'parent2_state': '', 'parent2_zip': '',
        'parent2_phone': '', 'parent2_email': '',
        'bill_to_parent': '1', 'bill_to_custom_name': '', 'bill_to_custom_addr': '',
        'bill_to_custom_city': '', 'bill_to_custom_state': '', 'bill_to_custom_zip': '',
        'intake_complete': 'YES', 'roi_complete': 'NO',
        'hourly_rate': '', 'notes': 'Works best with visual aids.',
    }
    sample_fill = PatternFill('solid', fgColor='F9FAFB')
    for col_idx, (key, *_) in enumerate(COLS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=SAMPLE.get(key, ''))
        cell.fill      = sample_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border    = thin_border

    ws2 = wb.create_sheet('Services Reference')
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 30
    ws2.cell(1, 1, 'Key (use in Services column)').font = Font(bold=True)
    ws2.cell(1, 2, 'Display Name').font = Font(bold=True)
    for i, (key, label) in enumerate(SERVICES, start=2):
        ws2.cell(i, 1, key)
        ws2.cell(i, 2, label)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name='cadence_students_template.xlsx')


@app.route('/settings/import-students', methods=['POST'])
def import_students():
    """Import students from an uploaded Excel file."""
    import openpyxl

    f = request.files.get('student_file')
    if not f or not f.filename:
        flash('No file selected.', 'error')
        return redirect(url_for('settings'))
    if not f.filename.lower().endswith(('.xlsx', '.xlsm')):
        flash('Please upload an .xlsx file.', 'error')
        return redirect(url_for('settings'))

    VALID_SERVICES = {k for k, _ in SERVICES}

    try:
        wb   = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws   = wb.active
        rows = list(ws.iter_rows(min_row=1, values_only=True))
    except Exception as e:
        flash(f'Could not read Excel file: {e}', 'error')
        return redirect(url_for('settings'))

    if not rows:
        flash('The file is empty.', 'error')
        return redirect(url_for('settings'))

    COL_KEYS = [
        'name', 'initials', 'email', 'phone', 'school', 'grade', 'birthday',
        'diagnosis', 'services', 'services_other', 'start_date', 'end_date',
        'test_date', 'parent1_name', 'parent2_name', 'parent_address',
        'parent_city', 'parent_state', 'parent_zip', 'parent2_address',
        'parent2_city', 'parent2_state', 'parent2_zip', 'parent2_phone',
        'parent2_email', 'bill_to_parent', 'bill_to_custom_name', 'bill_to_custom_addr',
        'bill_to_custom_city', 'bill_to_custom_state', 'bill_to_custom_zip',
        'intake_complete', 'roi_complete', 'hourly_rate', 'notes',
    ]

    header_map     = {}
    data_start_row = 0
    for row_idx, row in enumerate(rows):
        row_lower = [str(c).lower().replace(' ', '_').rstrip('_*') if c else '' for c in row]
        if 'name' in row_lower and 'initials' in row_lower:
            for col_i, val in enumerate(row_lower):
                clean = val.rstrip('_*')
                if clean in COL_KEYS:
                    header_map[clean] = col_i
            data_start_row = row_idx + 1
            # Skip the hint/notes row that immediately follows the header
            if data_start_row < len(rows):
                next_row     = rows[data_start_row]
                name_val     = str(next_row[header_map.get('name', 0)]     or '').strip()
                initials_val = str(next_row[header_map.get('initials', 1)] or '').strip()
                if len(name_val) > 60 or len(initials_val) > 5:
                    data_start_row += 1
            break

    if not header_map or 'name' not in header_map:
        flash('Could not find the header row. Make sure you are using the Cadence template.', 'error')
        return redirect(url_for('settings'))

    def _cell(row, key):
        idx = header_map.get(key)
        if idx is None or idx >= len(row):
            return ''
        v = row[idx]
        return str(v).strip() if v is not None else ''

    def _date_cell(row, key):
        v = (row[header_map[key]]
             if header_map.get(key) is not None and header_map[key] < len(row)
             else None)
        if v is None:
            return ''
        if hasattr(v, 'strftime'):
            return v.strftime('%Y-%m-%d')
        s = str(v).strip()
        for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y', '%d/%m/%Y'):
            try:
                return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
            except ValueError:
                pass
        return s

    conn     = get_db()
    existing = {r['name'].strip().lower()
                for r in conn.execute('SELECT name FROM clients').fetchall()}

    added = skipped = errors = 0
    error_msgs = []

    for row_idx, row in enumerate(rows[data_start_row:], start=data_start_row + 1):
        if not any(c for c in row if c is not None and str(c).strip()):
            continue

        name     = _cell(row, 'name')
        initials = _cell(row, 'initials').upper()
        if not name or not initials:
            errors += 1
            error_msgs.append(f'Row {row_idx}: skipped — Name and Initials are required.')
            continue

        if name.strip().lower() in existing:
            skipped += 1
            continue

        raw_svc   = _cell(row, 'services')
        svc_list  = [s.strip() for s in raw_svc.replace(';', ',').split(',') if s.strip()]
        svc_list  = [s for s in svc_list if s in VALID_SERVICES]
        svc_json  = json.dumps(svc_list) if svc_list else json.dumps([])

        intake = 1 if _cell(row, 'intake_complete').upper() == 'YES' else 0
        roi    = 1 if _cell(row, 'roi_complete').upper()    == 'YES' else 0

        try:
            hr_raw      = _cell(row, 'hourly_rate')
            hourly_rate = float(hr_raw.replace('$', '').replace(',', '')) if hr_raw else None
        except ValueError:
            hourly_rate = None

        raw_bill_to = _cell(row, 'bill_to_parent').lower()
        bill_to_parent = raw_bill_to if raw_bill_to in ('1', '2', 'custom') else '1'

        try:
            conn.execute('''
                INSERT INTO clients
                    (name, initials, email, phone, school, grade, birthday, diagnosis,
                     services, services_other, start_date, end_date, test_date,
                     parent1_name, parent2_name, parent_address, parent_city,
                     parent_state, parent_zip, parent2_address, parent2_city,
                     parent2_state, parent2_zip, parent2_phone, parent2_email,
                     bill_to_parent, bill_to_custom_name, bill_to_custom_addr,
                     bill_to_custom_city, bill_to_custom_state, bill_to_custom_zip,
                     intake_complete, roi_complete, notes, hourly_rate)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ''', (name, initials,
                  _cell(row, 'email'),      _cell(row, 'phone'),
                  _cell(row, 'school'),     _cell(row, 'grade'),
                  _date_cell(row, 'birthday'), _cell(row, 'diagnosis'),
                  svc_json, _cell(row, 'services_other'),
                  _date_cell(row, 'start_date'), _date_cell(row, 'end_date'),
                  _date_cell(row, 'test_date'),
                  _cell(row, 'parent1_name'), _cell(row, 'parent2_name'),
                  _cell(row, 'parent_address'),  _cell(row, 'parent_city'),
                  _cell(row, 'parent_state'),    _cell(row, 'parent_zip'),
                  _cell(row, 'parent2_address'), _cell(row, 'parent2_city'),
                  _cell(row, 'parent2_state'),   _cell(row, 'parent2_zip'),
                  _cell(row, 'parent2_phone'),   _cell(row, 'parent2_email'),
                  bill_to_parent,
                  _cell(row, 'bill_to_custom_name'), _cell(row, 'bill_to_custom_addr'),
                  _cell(row, 'bill_to_custom_city'),  _cell(row, 'bill_to_custom_state'),
                  _cell(row, 'bill_to_custom_zip'),
                  intake, roi, _cell(row, 'notes'), hourly_rate))
            existing.add(name.strip().lower())
            added += 1
        except Exception as e:
            errors += 1
            error_msgs.append(f'Row {row_idx} ({name}): {e}')

    conn.commit()
    conn.close()
    log.info('Student import: added=%d skipped=%d errors=%d', added, skipped, errors)

    parts = []
    if added:   parts.append(f'{added} student{"s" if added != 1 else ""} imported')
    if skipped: parts.append(f'{skipped} skipped (already exist)')
    if errors:  parts.append(f'{errors} error{"s" if errors != 1 else ""}')
    msg = ' · '.join(parts) if parts else 'Nothing to import.'
    cat = 'success' if added and not errors else ('warning' if added else 'error')
    flash(msg, cat)
    for em in error_msgs[:5]:
        flash(em, 'error')

    return redirect(url_for('settings'))


@app.route('/settings/storage/save', methods=['POST'])
def save_storage():
    new_db     = request.form.get('db_path',    '').strip()
    new_folder = request.form.get('pdf_folder', '').strip()

    current = _config.load()
    old_db  = current['db_path']

    if new_db and new_db != old_db:
        try:
            os.makedirs(os.path.dirname(os.path.abspath(new_db)), exist_ok=True)
            if os.path.exists(old_db) and not os.path.exists(new_db):
                shutil.copy2(old_db, new_db)
        except OSError as e:
            flash(f'Could not move database: {e}', 'error')
            return redirect(url_for('settings'))

    _config.save(new_db or old_db, new_folder or current['pdf_folder'])
    flash('Storage paths saved. Restart Cadence on both PCs for changes to take effect.', 'warning')
    return redirect(url_for('settings'))


@app.route('/settings/save', methods=['POST'])
def save_settings():
    hourly_rate_raw = request.form.get('hourly_rate', '').strip()
    if hourly_rate_raw:
        try:
            float(hourly_rate_raw)
        except ValueError:
            flash('Per-session rate must be a number.', 'error')
            return redirect(url_for('settings'))

    idle_raw = request.form.get('idle_timeout_minutes', '').strip()
    if idle_raw:
        try:
            int(idle_raw)
        except ValueError:
            flash('Idle timeout must be a whole number of minutes.', 'error')
            return redirect(url_for('settings'))

    for key in ('hourly_rate', 'business_name', 'business_title', 'business_email',
                'business_phone', 'business_address', 'business_city',
                'business_state', 'business_zip', 'venmo_handle', 'idle_timeout_minutes'):
        set_setting(key, request.form.get(key, ''))
    log.info('Settings saved')
    flash('Settings saved.', 'success')
    return redirect(url_for('settings'))


@app.route('/settings/calendars/save', methods=['POST'])
def save_calendars():
    default_cal = request.form.get('default_cal')
    conn        = get_db()
    all_cals    = conn.execute('SELECT id FROM calendars').fetchall()
    for cal in all_cals:
        enabled    = 1 if request.form.get(f'enabled_{cal["id"]}') else 0
        is_default = 1 if str(cal['id']) == default_cal else 0
        conn.execute('UPDATE calendars SET enabled=?, is_default=? WHERE id=?',
                     (enabled, is_default, cal['id']))
    conn.commit()
    conn.close()
    flash('Calendar preferences saved.', 'success')
    return redirect(url_for('settings'))


@app.route('/settings/graph/save', methods=['POST'])
def save_graph_settings():
    client_id = request.form.get('graph_client_id', '').strip()
    timezone  = request.form.get('graph_timezone',  '').strip()
    if client_id:
        set_setting('graph_client_id', client_id)
    if timezone:
        set_setting('graph_timezone', timezone)
    flash('Microsoft 365 settings saved.', 'success')
    return redirect(url_for('settings'))


@app.route('/settings/graph/connect', methods=['POST'])
def graph_connect():
    try:
        info = _graph_auth.start_device_flow()
        return jsonify(info)
    except RuntimeError as e:
        msg = str(e)
        if msg == 'no_client_id':
            msg = 'Enter and save your Azure App Client ID first.'
        return jsonify({'error': msg}), 400


@app.route('/settings/graph/poll')
def graph_auth_poll():
    return jsonify(_graph_auth.poll_auth())


@app.route('/settings/graph/disconnect', methods=['POST'])
def graph_disconnect():
    _graph_auth.disconnect()
    log.info('Disconnected from Microsoft 365')
    flash('Disconnected from Microsoft 365.', 'success')
    return redirect(url_for('settings'))


@app.route('/settings/calendars/refresh', methods=['POST'])
def refresh_calendars():
    try:
        cals = discover_calendars()
        upsert_calendars(cals)
        log.info('Refreshed calendars: found %d', len(cals))
        flash(f'Found {len(cals)} calendar(s) in Microsoft 365.', 'success')
    except RuntimeError as e:
        log.error('Calendar refresh failed: %s', e)
        flash(f'Could not read calendars: {e}', 'error')
    return redirect(url_for('settings'))


# Security routes (/settings/security/*) live in routes/auth.py
